from groq import Groq
from openpyxl import load_workbook
from datetime import datetime
import json
import re
import os

# ── SETUP ────────────────────────────────────────────────
client = Groq(api_key="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")
EXCEL_PATH = "stocksense_inventory.xlsx"

# ── JSON PARSER ──────────────────────────────────────────
def safe_parse(raw):
    if not raw:
        return None

    text = raw.strip()
    print(f"  Raw output: {repr(text)}")

    text = re.sub(r'^```json\s*', '', text)
    text = re.sub(r'^```\s*', '', text)
    text = re.sub(r'\s*```$', '', text)
    text = text.strip()

    match = re.search(r'\{.*?\}', text, re.DOTALL)
    if match:
        text = match.group()

    text = text.replace("'", '"')
    text = re.sub(r',\s*}', '}', text)

    try:
        return json.loads(text)
    except json.JSONDecodeError as e:
        print(f"  Parse failed: {e}")
        print(f"  Tried to parse: {repr(text)}")
        return None


# ── COMMAND PARSER ───────────────────────────────────────
def parse_command(transcript_text):
    print(f"\n{'='*55}")
    print(f"  Input: {transcript_text}")

    response = client.chat.completions.create(
        model="llama-3.3-70b-versatile",
        max_tokens=150,
        messages=[
            {
                "role": "system",
                "content": """
                                You are an inventory tracking assistant for a small business.

                                Extract exactly three things from voice commands:
                                - item: name of the item (string)
                                - quantity: how many taken or added (integer)
                                - action: either exactly "taken" or exactly "added"

                                Rules:
                                - Respond with JSON and nothing else
                                - No markdown, no backticks, no explanation whatsoever
                                - If quantity unclear use 1
                                - If item unclear use "unknown"
                                - If action unclear use "taken"
                                - "used", "took", "grabbed", "removed", "ran out" = taken
                                - "added", "delivery", "restocked", "put back", "got" = added

                                Your entire response must be exactly this format with no other text:
                                {"item": "name here", "quantity": 1, "action": "taken"}
                            """
            },
            {
                "role": "user",
                "content": transcript_text
            }
        ]
    )

    raw = response.choices[0].message.content
    parsed = safe_parse(raw)

    if parsed:
        parsed.setdefault("item", "unknown")
        parsed.setdefault("quantity", 1)
        parsed.setdefault("action", "taken")
        parsed["quantity"] = int(parsed.get("quantity", 1))
        parsed["action"] = str(parsed.get("action", "taken")).lower()
        if parsed["action"] not in ["taken", "added"]:
            parsed["action"] = "taken"

    print(f"  Parsed: {parsed}")
    return parsed


# ── EXCEL UPDATER ────────────────────────────────────────
def update_excel(item_name, quantity, action, raw_command):

    # Check file exists
    if not os.path.exists(EXCEL_PATH):
        print(f"  ERROR: Cannot find {EXCEL_PATH}")
        print(f"  Make sure the Excel file is in: {os.getcwd()}")
        return None

    wb = load_workbook(EXCEL_PATH)
    ws_inv = wb["Inventory"]
    ws_log = wb["Transaction Log"]

    # Find item row by partial name match
    found_row = None
    found_name = None
    for row in ws_inv.iter_rows(min_row=4, max_col=8):
        cell_name = row[1].value
        if cell_name and item_name.lower() in str(cell_name).lower():
            found_row = row[0].row
            found_name = cell_name
            break

    if not found_row:
        print(f"  WARNING: '{item_name}' not found in Excel — skipping update")
        return None

    # Get current values
    cur_qty = ws_inv.cell(row=found_row, column=4).value or 0
    min_qty = ws_inv.cell(row=found_row, column=5).value or 0
    item_id = ws_inv.cell(row=found_row, column=1).value

    # Calculate new quantity
    if action == "taken":
        new_qty = max(0, cur_qty - quantity)
    else:
        new_qty = cur_qty + quantity

    # Determine new status
    if new_qty == 0:
        new_status = "MISSING"
    elif new_qty <= min_qty:
        new_status = "LOW STOCK"
    else:
        new_status = "OK"

    # Update Inventory sheet
    ws_inv.cell(row=found_row, column=4).value = new_qty
    ws_inv.cell(row=found_row, column=8).value = new_status
    ws_inv.cell(row=found_row, column=9).value = datetime.now().strftime("%d %b %Y")

    # Append to Transaction Log
    next_row = ws_log.max_row + 1
    ws_log.cell(row=next_row, column=1).value = datetime.now().strftime("%d %b %Y %H:%M")
    ws_log.cell(row=next_row, column=2).value = item_id
    ws_log.cell(row=next_row, column=3).value = found_name
    ws_log.cell(row=next_row, column=4).value = action.upper()
    ws_log.cell(row=next_row, column=5).value = quantity
    ws_log.cell(row=next_row, column=6).value = raw_command
    ws_log.cell(row=next_row, column=7).value = new_qty

    wb.save(EXCEL_PATH)

    print(f"  Excel updated → {found_name}: {cur_qty} → {new_qty} ({new_status})")
    return new_status


# ── ALERT CHECKER ────────────────────────────────────────
def check_alert(item_name, status):
    if status == "MISSING":
        print(f"  ⚠️  ALERT: {item_name} is MISSING — needs immediate reorder")
    elif status == "LOW STOCK":
        print(f"  ⚠️  ALERT: {item_name} is LOW — reorder soon")
    elif status == "OK":
        print(f"  ✅  {item_name} stock updated successfully")


# ── MAIN PIPELINE ────────────────────────────────────────
def process(command):
    """Full pipeline: text → parse → excel update → alert"""
    result = parse_command(command)

    if not result or result["item"] == "unknown":
        print("  Could not identify item — skipping")
        return

    status = update_excel(
        item_name=result["item"],
        quantity=result["quantity"],
        action=result["action"],
        raw_command=command
    )

    if status:
        check_alert(result["item"], status)


# ── TEST COMMANDS — all items exist in Excel ──────────────
if __name__ == "__main__":

    print("=" * 55)
    print("  STOCKSENSE — VOICE COMMAND PIPELINE TEST")
    print(f"  Excel file: {os.path.abspath(EXCEL_PATH)}")
    print("=" * 55)

    test_commands = [
        # Taking items
        "used two bottles of developer 20vol",
        "took the last box of latex gloves large",
        "grabbed one pack of foil sheets",
        "removed three tubes of hair colour 4N",
        "ran out of cotton pads",

        # Adding items
        "delivery came, got 10 hand sanitiser",
        "restocked bleach powder, added 5 kg",
        "put back 6 boxes of disposable masks",
        "added 4 rolls of receipt paper",
        "got a new toner cartridge, adding 2",
    ]

    passed = 0
    failed = 0

    for cmd in test_commands:
        result = parse_command(cmd)
        if result and result["item"] != "unknown":
            status = update_excel(
                item_name=result["item"],
                quantity=result["quantity"],
                action=result["action"],
                raw_command=cmd
            )
            if status:
                check_alert(result["item"], status)
                passed += 1
            else:
                failed += 1
        else:
            failed += 1

    print(f"\n{'='*55}")
    print(f"  RESULTS: {passed} passed / {failed} failed out of {len(test_commands)} tests")
    print(f"  Open stocksense_inventory.xlsx to see all updates")
    print("=" * 55)
