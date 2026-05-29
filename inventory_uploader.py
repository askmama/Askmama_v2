"""
One-time script: reads stocksense_inventory.xlsx and uploads both sheets
('Inventory' and 'Transaction Log') as new tabs in your Google Sheet.

Run once:
    python inventory_uploader.py
"""
import os
from openpyxl import load_workbook
from dotenv import load_dotenv
import gspread
from google.oauth2.service_account import Credentials

load_dotenv()

SCOPES = [
    'https://www.googleapis.com/auth/spreadsheets',
    'https://www.googleapis.com/auth/drive'
]
EXCEL_PATH = "stocksense_inventory.xlsx"


def get_sheets_client():
    creds_file = os.getenv('GOOGLE_CREDENTIALS_FILE', 'credentials.json')
    creds = Credentials.from_service_account_file(creds_file, scopes=SCOPES)
    return gspread.authorize(creds)


def upload_worksheet(wb_sheet, gsheet, tab_name):
    """Upload all rows from an openpyxl sheet into a Google Sheet tab."""
    try:
        ws = gsheet.worksheet(tab_name)
        print(f"  Tab '{tab_name}' exists — clearing and overwriting...")
        ws.clear()
    except gspread.WorksheetNotFound:
        ws = gsheet.add_worksheet(title=tab_name, rows=1000, cols=26)
        print(f"  Created new tab '{tab_name}'")

    data = []
    for row in wb_sheet.iter_rows(values_only=True):
        data.append(['' if v is None else str(v) for v in row])

    if data:
        ws.update('A1', data)
        print(f"  Uploaded {len(data)} rows to '{tab_name}'")


def main():
    if not os.path.exists(EXCEL_PATH):
        print(f"ERROR: {EXCEL_PATH} not found in {os.getcwd()}")
        return

    sheet_name = os.getenv('GOOGLE_SHEET_NAME')
    if not sheet_name:
        print("ERROR: GOOGLE_SHEET_NAME not set in .env")
        return

    print(f"Reading {EXCEL_PATH}...")
    wb = load_workbook(EXCEL_PATH)
    print(f"  Sheets found: {wb.sheetnames}")

    print(f"\nConnecting to Google Sheet '{sheet_name}'...")
    client = get_sheets_client()
    gsheet = client.open(sheet_name)

    for tab_name in ["Inventory", "Transaction Log"]:
        if tab_name in wb.sheetnames:
            upload_worksheet(wb[tab_name], gsheet, tab_name)
        else:
            print(f"  NOTE: '{tab_name}' not found in Excel — skipping")

    print("\nDone! Open your Google Sheet to verify.")


if __name__ == "__main__":
    main()
